from __future__ import annotations

import calendar
import csv
import io
import json
import re
import sys
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "legal-admin-aliases.json"
BASE_URL = "https://www.mois.go.kr"
LIST_URL = (
    f"{BASE_URL}/frt/bbs/type001/commonSelectBoardList.do"
    "?bbsId=BBSMSTR_000000000052"
)
POPULATION_URL = "https://raw.githubusercontent.com/greatsong/modudata/main/data/population_latest.csv"

SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "Chrome/124 Safari/537.36 population-dashboard-updater"
        )
    }
)


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def get_html(url: str) -> str:
    response = SESSION.get(url, timeout=90, allow_redirects=True)
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    return response.text


def decode_flex(content: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace"), "utf-8-replace"


def load_population_index() -> tuple[dict[str, str], str, date, str]:
    response = SESSION.get(POPULATION_URL, timeout=180, allow_redirects=True)
    response.raise_for_status()
    text, encoding = decode_flex(response.content)
    reader = csv.reader(io.StringIO(text))
    headers = next(reader)

    period_match = None
    for header in headers:
        period_match = re.search(r"(\d{4})년\s*(\d{1,2})월_(?:계|남|여)_", clean(header))
        if period_match:
            break
    if not period_match:
        raise RuntimeError("population_latest.csv 헤더에서 기준 연월을 찾지 못했습니다.")

    year, month = map(int, period_match.groups())
    period = f"{year:04d}-{month:02d}"
    cutoff = date(year, month, calendar.monthrange(year, month)[1])

    code_to_name: dict[str, str] = {}
    for row in reader:
        if not row:
            continue
        raw = clean(row[0])
        match = re.search(r"\((\d{8,12})\)\s*$", raw)
        if not match:
            continue
        code = match.group(1)
        name = re.sub(r"\s*\(\d{8,12}\)\s*$", "", raw).strip()
        if name:
            code_to_name[code] = name

    if len(code_to_name) < 3000:
        raise RuntimeError(f"인구 CSV 코드 행이 비정상적으로 적습니다: {len(code_to_name)}개")
    return code_to_name, period, cutoff, encoding


def board_article_urls(max_pages: int = 8) -> list[str]:
    articles: dict[int, str] = {}
    for page in range(1, max_pages + 1):
        page_url = f"{LIST_URL}&pageIndex={page}"
        soup = BeautifulSoup(get_html(page_url), "html.parser")
        found_on_page = 0
        for anchor in soup.find_all("a", href=True):
            text = clean(anchor.get_text(" ", strip=True))
            href = anchor.get("href", "")
            if "commonSelectBoardArticle.do" not in href:
                continue
            if not (
                ("행정기관" in text and "관할구역" in text)
                or "주민등록주소코드" in text
            ):
                continue
            match = re.search(r"nttId=(\d+)", href)
            if not match:
                continue
            articles[int(match.group(1))] = urljoin(BASE_URL, href)
            found_on_page += 1
        if page > 1 and found_on_page == 0:
            break
    return [url for _, url in sorted(articles.items(), reverse=True)]


def source_from_article(article_url: str) -> tuple[date, str, str] | None:
    soup = BeautifulSoup(get_html(article_url), "html.parser")
    candidates: list[tuple[date, str]] = []
    for anchor in soup.find_all("a", href=True):
        text = clean(anchor.get_text(" ", strip=True))
        href = anchor.get("href", "")
        match = re.search(r"jscode(\d{8})\.zip", text, re.IGNORECASE)
        if not match or "말소" in text or "FileDown.do" not in href:
            continue
        yyyymmdd = match.group(1)
        effective = datetime.strptime(yyyymmdd, "%Y%m%d").date()
        candidates.append((effective, urljoin(BASE_URL, href)))
    if not candidates:
        return None
    effective, zip_url = max(candidates)
    return effective, article_url, zip_url


def discover_sources() -> list[tuple[date, str, str]]:
    sources: list[tuple[date, str, str]] = []
    errors: list[str] = []
    for article_url in board_article_urls():
        try:
            source = source_from_article(article_url)
            if source:
                sources.append(source)
        except Exception as error:
            errors.append(f"{article_url}: {error}")
    unique = {(item[0], item[2]): item for item in sources}
    ordered = sorted(unique.values(), key=lambda item: item[0], reverse=True)
    if not ordered:
        raise RuntimeError("행안부 게시판에서 주소코드 ZIP을 찾지 못했습니다. " + "; ".join(errors[:3]))
    return ordered


def discover_latest_source() -> tuple[str, str, str]:
    effective, article_url, zip_url = discover_sources()[0]
    return effective.isoformat(), article_url, zip_url


def discover_source_for_cutoff(cutoff: date) -> tuple[date, str, str]:
    candidates = [source for source in discover_sources() if source[0] <= cutoff]
    if not candidates:
        raise RuntimeError(f"인구 기준일 {cutoff.isoformat()} 이전의 공식 주소코드표를 찾지 못했습니다.")
    return max(candidates, key=lambda item: item[0])


def workbook_rows(content: bytes) -> Iterable[tuple[object, ...]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    yield from worksheet.iter_rows(values_only=True)


def find_header_indexes(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    variants = {
        "admin_code": ("행정동코드", "행정기관코드", "기관코드"),
        "sido": ("시도명",),
        "sigungu": ("시군구명",),
        "admin": ("읍면동명", "행정동명"),
        "legal_code": ("법정동코드", "관할법정동코드", "동리코드"),
        "legal": ("동리명", "법정동명"),
    }
    for row_index, row in enumerate(rows[:30]):
        normalized = [clean(value).replace(" ", "") for value in row]
        indexes: dict[str, int] = {}
        for canonical, names in variants.items():
            for name in names:
                target = name.replace(" ", "")
                if target in normalized:
                    indexes[canonical] = normalized.index(target)
                    break
        if len(indexes) == len(variants):
            return row_index, indexes
    raise RuntimeError("KiKmix에서 행정동코드·법정동코드·지역명 헤더를 찾지 못했습니다.")


def row_value(row: tuple[object, ...], index: int) -> str:
    if index >= len(row):
        return ""
    return clean(row[index])


def legal_full_name(sido: str, sigungu: str, admin: str, legal: str) -> str:
    parts = [sido]
    if sigungu:
        parts.append(sigungu)
    if legal.endswith("리") and admin.endswith(("읍", "면")):
        parts.extend([admin, legal])
    else:
        parts.append(legal)
    return clean(" ".join(parts))


def parse_kikmix_xlsx(
    content: bytes,
    population_by_code: dict[str, str],
) -> tuple[dict[str, set[str]], dict[str, object]]:
    rows = list(workbook_rows(content))
    header_row, indexes = find_header_indexes(rows)
    aliases: dict[str, set[str]] = defaultdict(set)
    unresolved_codes: dict[str, set[str]] = defaultdict(set)
    official_admin_codes: set[str] = set()
    matched_admin_codes: set[str] = set()

    for row in rows[header_row + 1 :]:
        if not row:
            continue
        sido = row_value(row, indexes["sido"])
        sigungu = row_value(row, indexes["sigungu"])
        admin = row_value(row, indexes["admin"])
        legal = row_value(row, indexes["legal"])
        admin_code = re.sub(r"\D", "", row_value(row, indexes["admin_code"]))
        if not sido or not admin or not legal or not admin_code:
            continue

        official_admin_codes.add(admin_code)
        legal_name = legal_full_name(sido, sigungu, admin, legal)
        population_name = population_by_code.get(admin_code)
        if population_name:
            aliases[legal_name].add(population_name)
            matched_admin_codes.add(admin_code)
        else:
            official_name = clean(" ".join(part for part in (sido, sigungu, admin) if part))
            unresolved_codes[admin_code].add(official_name)

    if len(aliases) < 10000:
        raise RuntimeError(f"코드 결합 후 법정동 검색 사전이 비정상적으로 적습니다: {len(aliases)}개")

    stats = {
        "official_admin_code_count": len(official_admin_codes),
        "matched_admin_code_count": len(matched_admin_codes),
        "unmatched_admin_code_count": len(official_admin_codes - matched_admin_codes),
        "population_code_count": len(population_by_code),
        "unmatched_admin_sample": [
            {"code": code, "official_names": sorted(names)}
            for code, names in sorted(unresolved_codes.items())[:200]
        ],
    }
    return aliases, stats


def select_kikmix_xlsx(archive: zipfile.ZipFile) -> str:
    candidates = [
        name
        for name in archive.namelist()
        if "kikmix" in name.lower()
        and name.lower().endswith(".xlsx")
        and "말소" not in name
        and not Path(name).name.startswith("~$")
    ]
    if not candidates:
        raise RuntimeError("압축파일에서 말소코드 제외 KiKmix 엑셀을 찾지 못했습니다.")
    candidates.sort(key=lambda name: (len(Path(name).parts), len(name)))
    return candidates[0]


def fetch_zip(url: str, article_url: str) -> bytes:
    response = SESSION.get(
        url,
        timeout=120,
        allow_redirects=True,
        headers={"Referer": article_url},
    )
    response.raise_for_status()
    content = response.content
    if not content.startswith(b"PK"):
        preview = content[:160].decode("utf-8", errors="replace")
        raise RuntimeError(f"행안부 응답이 ZIP이 아닙니다: {preview!r}")
    return content


def main() -> None:
    population_by_code, population_period, cutoff, population_encoding = load_population_index()
    effective, article_url, zip_url = discover_source_for_cutoff(cutoff)
    if len(sys.argv) > 1 and sys.argv[1].strip():
        zip_url = sys.argv[1].strip()

    content = fetch_zip(zip_url, article_url)
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xlsx_name = select_kikmix_xlsx(archive)
        aliases, coverage = parse_kikmix_xlsx(archive.read(xlsx_name), population_by_code)

    payload = {
        "updated": effective.isoformat(),
        "generated": date.today().isoformat(),
        "scope": "nationwide-official-code-matched",
        "population_period": population_period,
        "population_cutoff": cutoff.isoformat(),
        "population_encoding": population_encoding,
        "source": article_url,
        "download": zip_url,
        "coverage": coverage,
        "aliases": [
            {"legal": legal, "admins": sorted(admins)}
            for legal, admins in sorted(aliases.items())
        ],
    }
    OUTPUT.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
        newline="\n",
    )
    print(
        f"Wrote {OUTPUT.name}: {len(payload['aliases']):,} legal areas; "
        f"population={population_period}; official={effective.isoformat()}; "
        f"matched={coverage['matched_admin_code_count']:,}/"
        f"{coverage['official_admin_code_count']:,}; source={xlsx_name}"
    )


if __name__ == "__main__":
    main()
