from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import requests
from openpyxl import load_workbook

from update_legal_aliases import (
    clean,
    discover_latest_source,
    fetch_zip,
    select_kikmix_xlsx,
)

ROOT = Path(__file__).resolve().parents[1]
REPORT = ROOT / "region-coverage-report.json"
POPULATION_URL = "https://raw.githubusercontent.com/greatsong/modudata/main/data/population_latest.csv"


def decode_flex(content: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace"), "utf-8-replace"


def population_rows() -> tuple[list[dict[str, str]], str, list[str]]:
    response = requests.get(POPULATION_URL, timeout=180)
    response.raise_for_status()
    text, encoding = decode_flex(response.content)
    reader = csv.reader(io.StringIO(text))
    headers = next(reader)
    rows: list[dict[str, str]] = []
    for row in reader:
        if not row:
            continue
        raw = row[0].strip()
        match = re.search(r"\((\d{8,12})\)\s*$", raw)
        name = re.sub(r"\s*\(\d{8,12}\)\s*$", "", raw).strip()
        if not name:
            continue
        rows.append({"name": name, "code": match.group(1) if match else ""})
    return rows, encoding, headers


def workbook_rows(content: bytes) -> list[tuple[object, ...]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    return list(worksheet.iter_rows(values_only=True))


def find_header_row(rows: list[tuple[object, ...]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:30]):
        headers = [clean(value) for value in row]
        compact = [value.replace(" ", "") for value in headers]
        if "시도명" in compact and ("읍면동명" in compact or "행정동명" in compact):
            return index, headers
    raise RuntimeError("KiKmix header row not found")


def choose_column(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    compact = [header.replace(" ", "") for header in headers]
    for candidate in candidates:
        target = candidate.replace(" ", "")
        if target in compact:
            return compact.index(target)
    return None


def row_value(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return clean(row[index])


def main() -> None:
    population, population_encoding, population_headers = population_rows()
    population_by_code = {item["code"]: item["name"] for item in population if item["code"]}
    population_names = {item["name"] for item in population}

    effective_date, article_url, zip_url = discover_latest_source()
    archive_content = fetch_zip(zip_url, article_url)
    with zipfile.ZipFile(io.BytesIO(archive_content)) as archive:
        xlsx_name = select_kikmix_xlsx(archive)
        rows = workbook_rows(archive.read(xlsx_name))

    header_row, headers = find_header_row(rows)
    indexes = {
        "admin_code": choose_column(headers, ("행정기관코드", "행정동코드", "기관코드")),
        "legal_code": choose_column(headers, ("법정동코드", "관할법정동코드", "동리코드")),
        "sido": choose_column(headers, ("시도명",)),
        "sigungu": choose_column(headers, ("시군구명",)),
        "admin": choose_column(headers, ("읍면동명", "행정동명")),
        "legal": choose_column(headers, ("동리명", "법정동명")),
    }

    official_admin_names: set[str] = set()
    official_admin_codes: set[str] = set()
    legal_to_admin_names: dict[str, set[str]] = defaultdict(set)
    legal_to_admin_codes: dict[str, set[str]] = defaultdict(set)
    samples: dict[str, list[dict[str, str]]] = {key: [] for key in ("방배", "봉천", "역삼", "신림", "잠실", "상계", "화곡")}

    for row in rows[header_row + 1 :]:
        sido = row_value(row, indexes["sido"])
        sigungu = row_value(row, indexes["sigungu"])
        admin = row_value(row, indexes["admin"])
        legal = row_value(row, indexes["legal"])
        admin_code = re.sub(r"\D", "", row_value(row, indexes["admin_code"]))
        legal_code = re.sub(r"\D", "", row_value(row, indexes["legal_code"]))
        if not sido or not admin or not legal:
            continue
        admin_full = " ".join(part for part in (sido, sigungu, admin) if part).strip()
        legal_parts = [sido]
        if sigungu:
            legal_parts.append(sigungu)
        if legal.endswith("리") and admin.endswith(("읍", "면")):
            legal_parts.extend([admin, legal])
        else:
            legal_parts.append(legal)
        legal_full = " ".join(legal_parts).strip()
        official_admin_names.add(admin_full)
        if admin_code:
            official_admin_codes.add(admin_code)
        legal_to_admin_names[legal_full].add(admin_full)
        if admin_code:
            legal_to_admin_codes[legal_full].add(admin_code)
        for keyword in samples:
            if keyword in legal_full or keyword in admin_full:
                if len(samples[keyword]) < 30:
                    samples[keyword].append({
                        "legal": legal_full,
                        "admin": admin_full,
                        "admin_code": admin_code,
                        "legal_code": legal_code,
                        "population_name_by_code": population_by_code.get(admin_code, ""),
                    })

    unresolved_names = sorted(official_admin_names - population_names)
    official_codes_without_population = sorted(code for code in official_admin_codes if code not in population_by_code)
    population_codes_without_official = sorted(code for code in population_by_code if code not in official_admin_codes)

    mapped_by_name = 0
    mapped_by_code = 0
    aliases_without_any_population_match: list[dict[str, object]] = []
    aliases_with_partial_match: list[dict[str, object]] = []

    for legal, admin_names in sorted(legal_to_admin_names.items()):
        name_matches = sorted(name for name in admin_names if name in population_names)
        code_matches = sorted({population_by_code[code] for code in legal_to_admin_codes[legal] if code in population_by_code})
        if name_matches:
            mapped_by_name += 1
        if code_matches:
            mapped_by_code += 1
        expected_count = len(admin_names)
        all_matches = sorted(set(name_matches) | set(code_matches))
        if not all_matches:
            aliases_without_any_population_match.append({
                "legal": legal,
                "official_admins": sorted(admin_names),
                "official_codes": sorted(legal_to_admin_codes[legal]),
            })
        elif len(all_matches) < expected_count:
            aliases_with_partial_match.append({
                "legal": legal,
                "official_admins": sorted(admin_names),
                "matched_population_rows": all_matches,
            })

    leaf_counter = Counter(name.split()[-1] for name in population_names)
    report = {
        "population": {
            "url": POPULATION_URL,
            "encoding": population_encoding,
            "header_count": len(population_headers),
            "row_count": len(population),
            "coded_row_count": sum(bool(item["code"]) for item in population),
            "unique_code_count": len(population_by_code),
            "unique_name_count": len(population_names),
            "sample_names": sorted(population_names)[:20],
        },
        "official": {
            "effective_date": effective_date,
            "article_url": article_url,
            "zip_url": zip_url,
            "xlsx_name": xlsx_name,
            "headers": headers,
            "indexes": indexes,
            "admin_name_count": len(official_admin_names),
            "admin_code_count": len(official_admin_codes),
            "legal_area_count": len(legal_to_admin_names),
        },
        "coverage": {
            "legal_aliases_mapped_by_name": mapped_by_name,
            "legal_aliases_mapped_by_code": mapped_by_code,
            "official_admin_names_not_in_population_count": len(unresolved_names),
            "official_admin_codes_not_in_population_count": len(official_codes_without_population),
            "population_codes_not_in_official_count": len(population_codes_without_official),
            "legal_aliases_without_any_population_match_count": len(aliases_without_any_population_match),
            "legal_aliases_with_partial_match_count": len(aliases_with_partial_match),
            "official_admin_names_not_in_population_sample": unresolved_names[:200],
            "official_admin_codes_not_in_population_sample": official_codes_without_population[:200],
            "population_codes_not_in_official_sample": population_codes_without_official[:200],
            "legal_aliases_without_any_population_match_sample": aliases_without_any_population_match[:100],
            "legal_aliases_with_partial_match_sample": aliases_with_partial_match[:100],
        },
        "samples": samples,
        "population_leaf_duplicates": [
            {"leaf": leaf, "count": count}
            for leaf, count in leaf_counter.most_common(100)
            if count > 1
        ],
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8", newline="\n")
    print(json.dumps(report["coverage"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
